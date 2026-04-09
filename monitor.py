import os, json, smtplib, requests, datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

FILES = [
    "https://amc.com/ads.txt",
    "https://amc.com/app-ads.txt",
    "https://bbca.com/ads.txt",
    "https://bbca.com/app-ads.txt",
    "https://ifc.com/ads.txt",
    "https://ifc.com/app-ads.txt",
    "https://sundancetv.com/ads.txt",
    "https://sundancetv.com/app-ads.txt",
    "https://wetv.com/ads.txt",
    "https://wetv.com/app-ads.txt",
]

PARTNERS = {
    "amc.com":        "AMC Networks",
    "bbca.com":       "BBC America",
    "ifc.com":        "IFC",
    "sundancetv.com": "Sundance TV",
    "wetv.com":       "WE tv",
}

SNAPSHOTS_FILE  = "snapshots.json"
CHANGELOG_FILE  = "change_log.json"
EXCEL_FILE      = "adstxt_changes.xlsx"
HTML_FILE       = "index.html"
GMAIL_USER         = os.environ["GMAIL_USER"]
GMAIL_APP_PASSWORD = os.environ["GMAIL_APP_PASSWORD"]
NOTIFY_EMAIL       = os.environ["NOTIFY_EMAIL"]

HEADER_FILL = PatternFill("solid", start_color="1F3864")
CHANGED_FILL = PatternFill("solid", start_color="FCE4D6")
NEW_FILL     = PatternFill("solid", start_color="E2EFDA")
ERROR_FILL   = PatternFill("solid", start_color="FFF2CC")
ADD_FILL     = PatternFill("solid", start_color="E2EFDA")
DEL_FILL     = PatternFill("solid", start_color="FCE4D6")


def fetch(url):
    try:
        r = requests.get(url, timeout=15, headers={"User-Agent": "adstxt-monitor/1.0"})
        r.raise_for_status()
        return {"ok": True, "text": r.text.strip()}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def load_snapshots():
    if os.path.exists(SNAPSHOTS_FILE):
        with open(SNAPSHOTS_FILE) as f:
            return json.load(f)
    return {}


def save_snapshots(snapshots):
    with open(SNAPSHOTS_FILE, "w") as f:
        json.dump(snapshots, f, indent=2)


def load_changelog():
    if os.path.exists(CHANGELOG_FILE):
        with open(CHANGELOG_FILE) as f:
            return json.load(f)
    return {}


def save_changelog(changelog):
    with open(CHANGELOG_FILE, "w") as f:
        json.dump(changelog, f, indent=2)


def section_map(text):
    """Return dict mapping each line -> its nearest preceding # header."""
    mapping = {}
    current = "(no section)"
    for line in (text or "").splitlines():
        stripped = line.strip()
        if stripped.startswith("#"):
            current = stripped
        mapping[stripped] = current
    return mapping


def diff_lines(old, new):
    """Return added/removed lines each as list of {line, section} dicts."""
    old_lines = [l.strip() for l in (old or "").splitlines() if l.strip()]
    new_lines = [l.strip() for l in (new or "").splitlines() if l.strip()]
    old_set = set(old_lines)
    new_set = set(new_lines)

    old_smap = section_map(old)
    new_smap = section_map(new)

    added   = [{"line": l, "section": new_smap.get(l, "")} for l in new_lines if l not in old_set and not l.startswith("#")]
    removed = [{"line": l, "section": old_smap.get(l, "")} for l in old_lines if l not in new_set and not l.startswith("#")]
    return added, removed


def short(url):
    return url.replace("https://", "")


def domain(url):
    return url.replace("https://", "").split("/")[0]


def partner_name(url):
    return PARTNERS.get(domain(url), domain(url))


def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def auto_width(ws, min_w=12, max_w=60):
    for col in ws.columns:
        length = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, min_w), max_w)


def update_excel(results, snapshots, now_str):
    wb = load_workbook(EXCEL_FILE) if os.path.exists(EXCEL_FILE) else Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    if "Change Log" not in wb.sheetnames:
        ws_log = wb.create_sheet("Change Log", 0)
        ws_log.append(["Timestamp", "Partner", "File", "Status", "Lines Added", "Lines Removed", "Added (Section: Line)", "Removed (Section: Line)"])
        style_header(ws_log, 1, 8)
        ws_log.row_dimensions[1].height = 20
    else:
        ws_log = wb["Change Log"]

    for url, result in results.items():
        prev_text = snapshots.get(url, {}).get("text")
        if result["ok"]:
            if prev_text is None:
                status = "First snapshot"
                added, removed, fill = [], [], NEW_FILL
            else:
                added, removed = diff_lines(prev_text, result["text"])
                status = "Changed" if (added or removed) else "Unchanged"
                fill   = CHANGED_FILL if (added or removed) else None

            def fmt(items):
                return "\n".join(f'{x["section"]}: {x["line"]}' for x in items[:20]) or "-"

            row = [now_str, partner_name(url), short(url), status,
                   len(added), len(removed), fmt(added), fmt(removed)]
        else:
            status, fill = "Error", ERROR_FILL
            row = [now_str, partner_name(url), short(url), f"Error: {result['error']}", "", "", "", ""]

        ws_log.append(row)
        if fill:
            r = ws_log.max_row
            for c in range(1, 9):
                ws_log.cell(r, c).fill = fill
        ws_log.row_dimensions[ws_log.max_row].height = 15

    ws_log.column_dimensions["A"].width = 20
    ws_log.column_dimensions["B"].width = 18
    ws_log.column_dimensions["C"].width = 35
    ws_log.column_dimensions["D"].width = 16
    ws_log.column_dimensions["E"].width = 13
    ws_log.column_dimensions["F"].width = 15
    ws_log.column_dimensions["G"].width = 70
    ws_log.column_dimensions["H"].width = 70
    ws_log.freeze_panes = "A2"

    if "Current Snapshot" in wb.sheetnames:
        del wb["Current Snapshot"]
    ws_snap = wb.create_sheet("Current Snapshot")
    ws_snap.append(["Partner", "File", "Last Updated", "Line Count", "Status"])
    style_header(ws_snap, 1, 5)
    for url, result in results.items():
        if result["ok"]:
            ws_snap.append([partner_name(url), short(url), now_str, len(result["text"].splitlines()), "OK"])
        else:
            ws_snap.append([partner_name(url), short(url), now_str, "-", f"Error: {result['error']}"])
            ws_snap.cell(ws_snap.max_row, 5).fill = ERROR_FILL
    auto_width(ws_snap)
    ws_snap.freeze_panes = "A2"

    for url, result in results.items():
        if not result["ok"]:
            continue
        prev_text = snapshots.get(url, {}).get("text")
        if prev_text is None:
            continue
        added, removed = diff_lines(prev_text, result["text"])
        if not added and not removed:
            continue
        sname = short(url).replace("/", "_").replace(".", "_")[:31]
        if sname in wb.sheetnames:
            del wb[sname]
        ws_d = wb.create_sheet(sname)
        ws_d.append(["Type", "Section", "Line Content"])
        style_header(ws_d, 1, 3)
        for item in added:
            ws_d.append(["+ Added", item["section"], item["line"]])
            r = ws_d.max_row
            for c in [1, 2, 3]:
                ws_d.cell(r, c).fill = ADD_FILL
                ws_d.cell(r, c).font = Font(color="375623", name="Arial", size=9)
        for item in removed:
            ws_d.append(["- Removed", item["section"], item["line"]])
            r = ws_d.max_row
            for c in [1, 2, 3]:
                ws_d.cell(r, c).fill = DEL_FILL
                ws_d.cell(r, c).font = Font(color="843C0C", name="Arial", size=9)
        ws_d.column_dimensions["A"].width = 12
        ws_d.column_dimensions["B"].width = 25
        ws_d.column_dimensions["C"].width = 80
        ws_d.freeze_panes = "A2"

    wb.save(EXCEL_FILE)


def update_changelog(results, snapshots, now_str):
    changelog = load_changelog()
    for url, result in results.items():
        key = short(url)
        prev_text = snapshots.get(url, {}).get("text")
        if not result["ok"]:
            entry = {"date": now_str, "status": "error", "error": result["error"]}
        elif prev_text is None:
            entry = {"date": now_str, "status": "first_snapshot", "lines": len(result["text"].splitlines())}
        else:
            added, removed = diff_lines(prev_text, result["text"])
            if added or removed:
                entry = {"date": now_str, "status": "changed",
                         "added": added, "removed": removed,
                         "lines": len(result["text"].splitlines())}
            else:
                entry = {"date": now_str, "status": "unchanged",
                         "lines": len(result["text"].splitlines())}
        if key not in changelog:
            changelog[key] = []
        changelog[key].append(entry)
    save_changelog(changelog)
    return changelog


def esc(s):
    return (s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def render_diff_items(items, css_class, label_color):
    if not items:
        return ""
    # Group by section
    sections = {}
    for item in items:
        sec = item.get("section", "(no section)")
        sections.setdefault(sec, []).append(item["line"])
    html = ""
    for sec, lines in sections.items():
        html += f'<div class="section-header">{esc(sec)}</div>'
        for line in lines:
            html += f'<div class="line {css_class}">{esc(line)}</div>'
    return html


def generate_html(changelog, now_str):
    partner_files = {}
    for url in FILES:
        p = partner_name(url)
        partner_files.setdefault(p, [])
        partner_files[p].append(short(url))

    total_changes = sum(
        1 for entries in changelog.values()
        for e in entries if e["status"] == "changed"
    )

    partner_sections = ""
    for partner, files in sorted(partner_files.items()):
        file_blocks = ""
        for f in files:
            entries = changelog.get(f, [])
            if not entries:
                continue
            entries_desc = list(reversed(entries))
            last = entries_desc[0]
            last_status = last["status"]
            badge_map = {
                "changed":        '<span class="badge changed">Changed</span>',
                "unchanged":      '<span class="badge ok">Unchanged</span>',
                "error":          '<span class="badge error">Error</span>',
                "first_snapshot": '<span class="badge new">First snapshot</span>',
            }
            badge = badge_map.get(last_status, "")
            show_open = last_status in ("changed", "error")

            timeline = ""
            for e in entries_desc:
                if e["status"] == "changed":
                    added_html   = render_diff_items(e.get("added", []),   "add", "#375623")
                    removed_html = render_diff_items(e.get("removed", []), "del", "#843C0C")
                    diff_block = f'<div class="diff">{added_html}{removed_html}</div>' if (added_html or removed_html) else ""
                    na, nr = len(e.get("added", [])), len(e.get("removed", []))
                    meta = f'+{na} added &nbsp;·&nbsp; -{nr} removed &nbsp;·&nbsp; {e.get("lines","?")} lines total'
                    timeline += f'<div class="entry changed"><div class="entry-date">{e["date"]}</div><div class="entry-meta">{meta}</div>{diff_block}</div>'
                elif e["status"] == "error":
                    timeline += f'<div class="entry error"><div class="entry-date">{e["date"]}</div><div class="entry-meta">Error: {esc(e.get("error",""))}</div></div>'
                elif e["status"] == "first_snapshot":
                    timeline += f'<div class="entry new"><div class="entry-date">{e["date"]}</div><div class="entry-meta">First snapshot saved &nbsp;·&nbsp; {e.get("lines","?")} lines</div></div>'
                else:
                    timeline += f'<div class="entry ok"><div class="entry-date">{e["date"]}</div><div class="entry-meta">No changes &nbsp;·&nbsp; {e.get("lines","?")} lines</div></div>'

            change_count = sum(1 for e in entries if e["status"] == "changed")
            count_tag = f'<span class="count">{change_count} change{"s" if change_count!=1 else ""}</span>' if change_count else ""
            open_attr = "open" if show_open else ""
            file_blocks += f'''
            <details class="file-block" {open_attr}>
              <summary><span class="file-name">{f}</span>{badge}{count_tag}</summary>
              <div class="timeline">{timeline}</div>
            </details>'''

        partner_sections += f'''
        <section class="partner">
          <h2>{partner}</h2>
          {file_blocks}
        </section>'''

    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>ads.txt Monitor</title>
<style>
  *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0 }}
  body {{ font-family: Arial, sans-serif; background: #f4f6f9; color: #1a1a2e; font-size: 14px; line-height: 1.5 }}
  header {{ background: #1F3864; color: #fff; padding: 24px 32px }}
  header h1 {{ font-size: 22px; font-weight: 600; margin-bottom: 4px }}
  header p {{ color: #BDD7EE; font-size: 13px }}
  .stats {{ display: flex; gap: 16px; padding: 20px 32px; background: #fff; border-bottom: 1px solid #e0e0e0; flex-wrap: wrap }}
  .stat {{ background: #f4f6f9; border-radius: 8px; padding: 12px 20px; min-width: 140px }}
  .stat-val {{ font-size: 24px; font-weight: 700; color: #1F3864 }}
  .stat-lbl {{ font-size: 12px; color: #666; margin-top: 2px }}
  main {{ max-width: 960px; margin: 28px auto; padding: 0 20px 60px }}
  .partner {{ background: #fff; border-radius: 10px; border: 1px solid #e0e0e0; margin-bottom: 24px; overflow: hidden }}
  .partner h2 {{ font-size: 16px; font-weight: 600; padding: 16px 20px; background: #f0f4fa; border-bottom: 1px solid #e0e0e0; color: #1F3864 }}
  .file-block {{ border-bottom: 1px solid #f0f0f0 }}
  .file-block:last-child {{ border-bottom: none }}
  .file-block summary {{ display: flex; align-items: center; gap: 10px; padding: 13px 20px; cursor: pointer; list-style: none; user-select: none }}
  .file-block summary:hover {{ background: #fafbfc }}
  .file-name {{ font-family: monospace; font-size: 13px; flex: 1; color: #333 }}
  .badge {{ font-size: 11px; padding: 2px 10px; border-radius: 20px; font-weight: 600 }}
  .badge.changed {{ background: #FCE4D6; color: #843C0C }}
  .badge.ok {{ background: #f0f0f0; color: #666 }}
  .badge.error {{ background: #FFF2CC; color: #7B6000 }}
  .badge.new {{ background: #E2EFDA; color: #375623 }}
  .count {{ font-size: 11px; color: #999; margin-left: 4px }}
  .timeline {{ padding: 0 20px 16px }}
  .entry {{ margin-top: 12px; border-radius: 6px; padding: 10px 14px; border-left: 3px solid #ddd }}
  .entry.changed {{ border-left-color: #E24B4A; background: #fff9f9 }}
  .entry.ok {{ border-left-color: #ccc; background: #fafafa }}
  .entry.error {{ border-left-color: #f0ad4e; background: #fffdf0 }}
  .entry.new {{ border-left-color: #1D9E75; background: #f6fffa }}
  .entry-date {{ font-size: 12px; color: #999; margin-bottom: 3px }}
  .entry-meta {{ font-size: 13px; color: #555 }}
  .diff {{ margin-top: 10px; border-radius: 4px; overflow: hidden; font-family: monospace; font-size: 12px }}
  .section-header {{ background: #e8edf5; color: #1F3864; padding: 4px 10px; font-weight: 700; font-size: 11px; letter-spacing: 0.03em; margin-top: 4px }}
  .line {{ padding: 2px 10px; white-space: pre-wrap; word-break: break-all }}
  .line.add {{ background: #E2EFDA; color: #375623 }}
  .line.del {{ background: #FCE4D6; color: #843C0C }}
  details[open] summary {{ background: #f7f9fc }}
</style>
</head>
<body>
<header>
  <h1>ads.txt Monitor</h1>
  <p>Last updated: {now_str} &nbsp;·&nbsp; Tracking {len(FILES)} files across {len(PARTNERS)} partners &nbsp;·&nbsp; <a href="competitors.html" style="color:#BDD7EE">Competitor Monitor →</a></p>
</header>
<div class="stats">
  <div class="stat"><div class="stat-val">{len(PARTNERS)}</div><div class="stat-lbl">Partners</div></div>
  <div class="stat"><div class="stat-val">{len(FILES)}</div><div class="stat-lbl">Files monitored</div></div>
  <div class="stat"><div class="stat-val">{total_changes}</div><div class="stat-lbl">Total changes detected</div></div>
</div>
<main>{partner_sections}</main>
</body>
</html>'''

    with open(HTML_FILE, "w", encoding="utf-8") as f:
        f.write(html)


def build_email_html(results, snapshots, now_str):
    changes, errors, unchanged, fresh = [], [], [], []
    for url, result in results.items():
        prev = snapshots.get(url, {}).get("text")
        if not result["ok"]:
            errors.append((url, result["error"]))
        elif prev is None:
            fresh.append(url)
        else:
            added, removed = diff_lines(prev, result["text"])
            if added or removed:
                changes.append((url, added, removed))
            else:
                unchanged.append(url)

    def badge(text, color, bg):
        return f'<span style="background:{bg};color:{color};padding:2px 8px;border-radius:12px;font-size:12px;font-weight:600">{text}</span>'

    def fmt_items(items):
        sections = {}
        for item in items:
            sec = item.get("section", "(no section)")
            sections.setdefault(sec, []).append(item["line"])
        out = ""
        for sec, lines in sections.items():
            out += f'<div style="background:#dde4f0;color:#1F3864;font-weight:700;font-size:11px;padding:3px 8px;margin-top:4px">{esc(sec)}</div>'
            for line in lines[:20]:
                out += f'<div style="padding:1px 8px">{esc(line)}</div>'
        return out

    rows = ""
    for url, added, removed in changes:
        rows += (f'<tr><td style="padding:10px 12px;font-size:12px;color:#555">{partner_name(url)}</td>'
                 f'<td style="padding:10px 12px;font-family:monospace;font-size:11px">{short(url)}</td>'
                 f'<td style="padding:10px 12px;text-align:center">{badge("CHANGED","#843C0C","#FCE4D6")}</td>'
                 f'<td style="padding:10px 12px;text-align:center;color:#375623">+{len(added)}</td>'
                 f'<td style="padding:10px 12px;text-align:center;color:#843C0C">-{len(removed)}</td></tr>')
        if added:
            rows += (f'<tr><td colspan="5" style="padding:4px 12px 8px">'
                     f'<details><summary style="font-size:12px;cursor:pointer;color:#375623">Show {len(added)} added lines</summary>'
                     f'<div style="background:#E2EFDA;font-family:monospace;font-size:11px;margin-top:6px;border-radius:4px;overflow:hidden">{fmt_items(added)}</div>'
                     f'</details></td></tr>')
        if removed:
            rows += (f'<tr><td colspan="5" style="padding:4px 12px 8px">'
                     f'<details><summary style="font-size:12px;cursor:pointer;color:#843C0C">Show {len(removed)} removed lines</summary>'
                     f'<div style="background:#FCE4D6;font-family:monospace;font-size:11px;margin-top:6px;border-radius:4px;overflow:hidden">{fmt_items(removed)}</div>'
                     f'</details></td></tr>')
    for url, err in errors:
        rows += (f'<tr><td style="padding:10px 12px;font-size:12px;color:#555">{partner_name(url)}</td>'
                 f'<td style="padding:10px 12px;font-family:monospace;font-size:11px">{short(url)}</td>'
                 f'<td style="padding:10px 12px;text-align:center">{badge("ERROR","#7B6000","#FFF2CC")}</td>'
                 f'<td colspan="2" style="padding:10px 12px;font-size:12px;color:#7B6000">{err}</td></tr>')
    for url in unchanged:
        rows += (f'<tr style="color:#aaa"><td style="padding:8px 12px;font-size:12px">{partner_name(url)}</td>'
                 f'<td style="padding:8px 12px;font-family:monospace;font-size:11px">{short(url)}</td>'
                 f'<td style="padding:8px 12px;text-align:center">{badge("OK","#555","#eee")}</td><td colspan="2"></td></tr>')
    for url in fresh:
        rows += (f'<tr><td style="padding:8px 12px;font-size:12px;color:#555">{partner_name(url)}</td>'
                 f'<td style="padding:8px 12px;font-family:monospace;font-size:11px">{short(url)}</td>'
                 f'<td style="padding:8px 12px;text-align:center">{badge("NEW","#1F3864","#BDD7EE")}</td>'
                 f'<td colspan="2" style="font-size:12px;color:#555">First snapshot saved</td></tr>')

    subject_tag = f"🔴 {len(changes)} change(s)" if changes else ("⚠️ errors" if errors else "✅ no changes")
    summary = f"{len(changes)} changed · {len(errors)} errors · {len(unchanged)} unchanged · {len(fresh)} new snapshots"

    html = (f'<!DOCTYPE html><html><head><meta charset="utf-8"></head>'
            f'<body style="font-family:Arial,sans-serif;background:#f5f5f5;margin:0;padding:20px">'
            f'<div style="max-width:760px;margin:auto;background:#fff;border-radius:8px;overflow:hidden;border:1px solid #ddd">'
            f'<div style="background:#1F3864;padding:20px 24px">'
            f'<h1 style="color:#fff;margin:0;font-size:18px">ads.txt Monitor Report</h1>'
            f'<p style="color:#BDD7EE;margin:6px 0 0;font-size:13px">{now_str} · {summary}</p></div>'
            f'<table style="width:100%;border-collapse:collapse">'
            f'<thead><tr style="background:#F2F2F2;font-size:12px;color:#555">'
            f'<th style="padding:10px 12px;text-align:left">Partner</th>'
            f'<th style="padding:10px 12px;text-align:left">File</th>'
            f'<th style="padding:10px 12px">Status</th>'
            f'<th style="padding:10px 12px">Added</th>'
            f'<th style="padding:10px 12px">Removed</th></tr></thead>'
            f'<tbody>{rows}</tbody></table>'
            f'<div style="padding:16px 24px;font-size:12px;color:#888;border-top:1px solid #eee">'
            f'Full diff details are attached in the Excel file. Sent daily by your GitHub Actions monitor.'
            f'</div></div></body></html>')

    return subject_tag, html


def send_email(subject_tag, html_body, now_str):
    subject = f"[ads.txt Monitor] {subject_tag} - {now_str}"
    msg = MIMEMultipart("mixed")
    msg["From"] = GMAIL_USER
    msg["To"] = NOTIFY_EMAIL
    msg["Subject"] = subject
    msg.attach(MIMEText(html_body, "html"))
    with open(EXCEL_FILE, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{EXCEL_FILE}"')
        msg.attach(part)
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(GMAIL_USER, GMAIL_APP_PASSWORD)
        server.sendmail(GMAIL_USER, NOTIFY_EMAIL, msg.as_string())


def generate_amc_html(changelog, now_str):
    """Flat, fully-expanded page for AMC Networks files — no <details>, AI-readable."""
    amc_files = ["amc.com/ads.txt", "amc.com/app-ads.txt"]
    CSS = """
  *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0 }
  body { font-family: Arial, sans-serif; background: #f4f6f9; color: #1a1a2e; font-size: 14px; line-height: 1.5 }
  header { background: #1F3864; color: #fff; padding: 24px 32px }
  header h1 { font-size: 22px; font-weight: 600; margin-bottom: 4px }
  header p { color: #BDD7EE; font-size: 13px }
  main { max-width: 960px; margin: 28px auto; padding: 0 20px 60px }
  .file-block { background: #fff; border-radius: 10px; border: 1px solid #e0e0e0; margin-bottom: 28px; overflow: hidden }
  .file-title { font-size: 15px; font-weight: 600; padding: 14px 20px; background: #f0f4fa; border-bottom: 1px solid #e0e0e0; color: #1F3864; font-family: monospace }
  .entry { margin: 12px 16px; border-radius: 6px; padding: 10px 14px; border-left: 3px solid #ddd }
  .entry.changed { border-left-color: #E24B4A; background: #fff9f9 }
  .entry.ok { border-left-color: #ccc; background: #fafafa }
  .entry.error { border-left-color: #f0ad4e; background: #fffdf0 }
  .entry.new { border-left-color: #1D9E75; background: #f6fffa }
  .entry-date { font-size: 12px; color: #999; margin-bottom: 3px }
  .entry-meta { font-size: 13px; color: #555; margin-bottom: 6px }
  .diff { border-radius: 4px; overflow: hidden; font-family: monospace; font-size: 12px }
  .section-header { background: #e8edf5; color: #1F3864; padding: 4px 10px; font-weight: 700; font-size: 11px; margin-top: 4px }
  .line { padding: 2px 10px; white-space: pre-wrap; word-break: break-all }
  .line.add { background: #E2EFDA; color: #375623 }
  .line.del { background: #FCE4D6; color: #843C0C }
"""
    file_blocks = ""
    for f in amc_files:
        entries = list(reversed(changelog.get(f, [])))
        timeline = ""
        for e in entries:
            if e["status"] == "changed":
                added_html   = render_diff_items(e.get("added", []),   "add", "#375623")
                removed_html = render_diff_items(e.get("removed", []), "del", "#843C0C")
                diff_block = f'<div class="diff">{added_html}{removed_html}</div>'
                na, nr = len(e.get("added", [])), len(e.get("removed", []))
                meta = f'+{na} added &nbsp;&middot;&nbsp; -{nr} removed &nbsp;&middot;&nbsp; {e.get("lines","?")} lines total'
                timeline += f'<div class="entry changed"><div class="entry-date">{e["date"]}</div><div class="entry-meta">{meta}</div>{diff_block}</div>'
            elif e["status"] == "error":
                timeline += f'<div class="entry error"><div class="entry-date">{e["date"]}</div><div class="entry-meta">Error: {esc(e.get("error",""))}</div></div>'
            elif e["status"] == "first_snapshot":
                timeline += f'<div class="entry new"><div class="entry-date">{e["date"]}</div><div class="entry-meta">First snapshot &nbsp;&middot;&nbsp; {e.get("lines","?")} lines</div></div>'
            else:
                timeline += f'<div class="entry ok"><div class="entry-date">{e["date"]}</div><div class="entry-meta">No changes &nbsp;&middot;&nbsp; {e.get("lines","?")} lines</div></div>'
        file_blocks += f'<div class="file-block"><div class="file-title">{f}</div>{timeline}</div>'

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>AMC Networks — ads.txt Change Log</title>
<style>{CSS}</style>
</head>
<body>
<header>
  <h1>AMC Networks — ads.txt Change Log</h1>
  <p>Last updated: {now_str} &nbsp;·&nbsp; All entries fully expanded &nbsp;·&nbsp; <a href="index.html" style="color:#BDD7EE">View all partners</a></p>
</header>
<main>{file_blocks}</main>
</body>
</html>"""

    with open("amc.html", "w", encoding="utf-8") as f:
        f.write(html)


def main():
    now_str = datetime.datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    snapshots = load_snapshots()

    print(f"Checking {len(FILES)} files...")
    results = {}
    for url in FILES:
        print(f"  {short(url)}", end=" ", flush=True)
        results[url] = fetch(url)
        print("ok" if results[url]["ok"] else "error")

    update_excel(results, snapshots, now_str)
    print(f"Excel updated -> {EXCEL_FILE}")

    changelog = update_changelog(results, snapshots, now_str)
    generate_html(changelog, now_str)
    generate_amc_html(changelog, now_str)
    print(f"HTML log updated -> {HTML_FILE} + amc.html")

    subject_tag, html = build_email_html(results, snapshots, now_str)
    send_email(subject_tag, html, now_str)
    print(f"Email sent -> {NOTIFY_EMAIL}")

    for url, result in results.items():
        if result["ok"]:
            snapshots[url] = {"text": result["text"], "updated": now_str}
    save_snapshots(snapshots)
    print("Snapshots saved.")


if __name__ == "__main__":
    main()
